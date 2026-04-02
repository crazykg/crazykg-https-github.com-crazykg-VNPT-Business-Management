#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import json
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


IMPORT_HEADERS = [
    'Mã NV',
    'Họ tên',
    'Ngày sinh',
    'Loại đảng viên',
    'Ngày vào Đảng dự bị',
    'Ngày chính thức',
    'Quê quán',
    'Dân tộc',
    'Tôn giáo',
    'Trình độ chuyên môn',
    'LLCT',
    'Số thẻ Đảng',
    'Ngày cấp thẻ',
    'Ghi chú',
]

REFERENCE_HEADERS = ['Mã NV', 'Họ tên', 'Ngày sinh', 'Phòng ban', 'Chức vụ']
REPORT_HEADERS = [
    'Dòng nguồn',
    'Họ tên trong file',
    'Ngày sinh trong file',
    'Gợi ý Mã NV',
    'Họ tên hệ thống',
    'Ngày sinh hệ thống',
    'Phòng ban',
    'Chức vụ',
    'Kết quả map',
    'Ghi chú',
]
MANUAL_REVIEW_HEADERS = [
    'Dòng nguồn',
    'Họ tên trong file',
    'Ngày sinh trong file',
    'Loại đảng viên',
    'Nhóm xử lý',
    'Gợi ý đối chiếu',
    'Ghi chú',
]


@dataclass
class EmployeeRef:
    user_code: str
    full_name: str
    date_of_birth: str | None
    dept_code: str
    dept_name: str
    pos_name: str

    @property
    def normalized_name(self) -> str:
        return normalize_name(self.full_name)


@dataclass
class PartyRow:
    source_row: int
    full_name: str
    date_of_birth: str | None
    party_member_status: str
    probationary_join_date: str | None
    official_join_date: str | None
    hometown: str | None
    ethnicity: str | None
    religion: str | None
    professional_qualification: str | None
    political_theory_level: str | None
    party_card_number: str | None
    party_card_issue_date: str | None
    notes: str | None

    @property
    def normalized_name(self) -> str:
        return normalize_name(self.full_name)


@dataclass
class MappingResult:
    party_row: PartyRow
    status: str
    match: EmployeeRef | None
    note: str
    suggestions: list[EmployeeRef]

    @property
    def auto_fill_code(self) -> str:
        if self.match is None:
            return ''
        if self.status in {'EXACT_NAME_DOB', 'NAME_ONLY'}:
            return self.match.user_code
        return ''


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace('\n', ' ').strip()
    return ' '.join(text.split()) or None


def normalize_name(value: str | None) -> str:
    if not value:
        return ''
    normalized = unicodedata.normalize('NFD', value)
    normalized = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
    normalized = normalized.replace('đ', 'd').replace('Đ', 'D')
    normalized = ' '.join(normalized.lower().split())
    return normalized


def format_date_value(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')

    text = normalize_text(value)
    if not text:
        return None

    for pattern in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, pattern).strftime('%Y-%m-%d')
        except ValueError:
            continue

    return None


def load_employee_refs(tsv_path: Path) -> list[EmployeeRef]:
    rows: list[EmployeeRef] = []
    with tsv_path.open('r', encoding='utf-8') as handle:
        reader = csv.DictReader(handle, delimiter='\t')
        for item in reader:
            rows.append(
                EmployeeRef(
                    user_code=normalize_text(item.get('user_code')) or '',
                    full_name=normalize_text(item.get('full_name')) or '',
                    date_of_birth=normalize_text(item.get('date_of_birth')),
                    dept_code=normalize_text(item.get('dept_code')) or '',
                    dept_name=normalize_text(item.get('dept_name')) or '',
                    pos_name=normalize_text(item.get('pos_name')) or '',
                )
            )
    return rows


def load_party_rows(source_path: Path) -> list[PartyRow]:
    workbook = load_workbook(source_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[PartyRow] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=6, values_only=True), start=6):
        values = list(row[:14])
        if not any(value is not None and str(value).strip() for value in values):
            continue

        male_dob = format_date_value(values[2] if len(values) > 2 else None)
        female_dob = format_date_value(values[3] if len(values) > 3 else None)
        probationary = format_date_value(values[7] if len(values) > 7 else None)
        official = format_date_value(values[8] if len(values) > 8 else None)
        card_issue = format_date_value(values[12] if len(values) > 12 else None)
        rows.append(
            PartyRow(
                source_row=row_number,
                full_name=normalize_text(values[1] if len(values) > 1 else None) or '',
                date_of_birth=male_dob or female_dob,
                party_member_status='OFFICIAL' if official else 'PROBATIONARY',
                probationary_join_date=probationary,
                official_join_date=official,
                hometown=normalize_text(values[6] if len(values) > 6 else None),
                ethnicity=normalize_text(values[4] if len(values) > 4 else None),
                religion=normalize_text(values[5] if len(values) > 5 else None),
                professional_qualification=normalize_text(values[9] if len(values) > 9 else None),
                political_theory_level=normalize_text(values[10] if len(values) > 10 else None),
                party_card_number=normalize_text(values[11] if len(values) > 11 else None),
                party_card_issue_date=card_issue,
                notes=normalize_text(values[13] if len(values) > 13 else None),
            )
        )

    return rows


def build_name_index(employees: Iterable[EmployeeRef]) -> dict[str, list[EmployeeRef]]:
    index: dict[str, list[EmployeeRef]] = {}
    for employee in employees:
        index.setdefault(employee.normalized_name, []).append(employee)
    return index


def build_department_label(employee: EmployeeRef | None) -> str:
    if employee is None:
        return ''
    return ' - '.join(part for part in [employee.dept_code, employee.dept_name] if part)


def extract_birth_year(date_value: str | None) -> str | None:
    if not date_value:
        return None
    try:
        return str(datetime.strptime(date_value, '%Y-%m-%d').year)
    except ValueError:
        return None


def is_note_row(party_row: PartyRow) -> bool:
    return party_row.normalized_name.startswith('luu y:')


def match_score(party_row: PartyRow, employee: EmployeeRef) -> float:
    name_ratio = SequenceMatcher(None, party_row.normalized_name, employee.normalized_name).ratio()
    party_tokens = set(party_row.normalized_name.split())
    employee_tokens = set(employee.normalized_name.split())
    token_overlap = len(party_tokens & employee_tokens) / max(1, len(party_tokens | employee_tokens))
    score = max(name_ratio, token_overlap)

    if party_row.date_of_birth and employee.date_of_birth:
        if party_row.date_of_birth == employee.date_of_birth:
            score += 0.35
        elif extract_birth_year(party_row.date_of_birth) == extract_birth_year(employee.date_of_birth):
            score += 0.08

    party_parts = party_row.normalized_name.split()
    employee_parts = employee.normalized_name.split()
    if party_parts and employee_parts and party_parts[-1] == employee_parts[-1]:
        score += 0.08
    if party_parts and employee_parts and party_parts[0] == employee_parts[0]:
        score += 0.04

    return score


def find_review_candidates(
    party_row: PartyRow,
    employees: list[EmployeeRef],
    limit: int = 3,
) -> list[EmployeeRef]:
    scored: list[tuple[float, EmployeeRef]] = []
    for employee in employees:
        score = match_score(party_row, employee)
        if score < 0.45:
            continue
        scored.append((score, employee))

    scored.sort(key=lambda item: (-item[0], item[1].full_name))
    suggestions: list[EmployeeRef] = []
    seen_codes: set[str] = set()
    for _, employee in scored:
        if employee.user_code in seen_codes:
            continue
        seen_codes.add(employee.user_code)
        suggestions.append(employee)
        if len(suggestions) >= limit:
            break
    return suggestions


def format_suggestions(suggestions: list[EmployeeRef]) -> str:
    if not suggestions:
        return ''
    parts = []
    for employee in suggestions:
        segment = f'{employee.user_code} - {employee.full_name}'
        if employee.date_of_birth:
            segment += f' ({employee.date_of_birth})'
        department = build_department_label(employee)
        if department:
            segment += f' [{department}]'
        parts.append(segment)
    return ' | '.join(parts)


def choose_match(party_row: PartyRow, candidates: list[EmployeeRef]) -> tuple[str, EmployeeRef | None, str]:
    if is_note_row(party_row):
        return 'NOT_FOUND', None, 'Dòng ghi chú trong file nguồn, bỏ qua khi import.'

    if not candidates:
        return 'NOT_FOUND', None, 'Không tìm thấy nhân sự trùng tên trong hệ thống.'

    same_dob = [item for item in candidates if item.date_of_birth and item.date_of_birth == party_row.date_of_birth]
    if len(same_dob) == 1:
        return 'EXACT_NAME_DOB', same_dob[0], 'Khớp duy nhất theo họ tên và ngày sinh.'
    if len(same_dob) > 1:
        return 'AMBIGUOUS_NAME_DOB', None, 'Có nhiều nhân sự trùng họ tên và ngày sinh.'

    if len(candidates) == 1:
        candidate = candidates[0]
        if party_row.date_of_birth and candidate.date_of_birth and candidate.date_of_birth != party_row.date_of_birth:
            note = f'Lệch ngày sinh: file={party_row.date_of_birth}, hệ thống={candidate.date_of_birth}.'
        elif party_row.date_of_birth and not candidate.date_of_birth:
            note = 'Khớp tên duy nhất nhưng hệ thống chưa có ngày sinh để đối chiếu.'
        else:
            note = 'Khớp tên duy nhất.'
        return 'NAME_ONLY', candidate, note

    return 'AMBIGUOUS_NAME', None, f'Tìm thấy {len(candidates)} nhân sự trùng tên, cần rà soát thủ công.'


def autosize_columns(worksheet) -> None:
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def write_workbook(
    mapping_results: list[MappingResult],
    employees: list[EmployeeRef],
    output_path: Path,
    report_rows: list[list[object]],
    manual_review_rows: list[list[object]],
) -> None:
    workbook = Workbook()

    dang_vien = workbook.active
    dang_vien.title = 'DangVien'
    dang_vien.append(IMPORT_HEADERS)
    for result in mapping_results:
        row = result.party_row
        dang_vien.append([
            result.auto_fill_code,
            row.full_name,
            row.date_of_birth or '',
            row.party_member_status,
            row.probationary_join_date or '',
            row.official_join_date or '',
            row.hometown or '',
            row.ethnicity or '',
            row.religion or '',
            row.professional_qualification or '',
            row.political_theory_level or '',
            row.party_card_number or '',
            row.party_card_issue_date or '',
            row.notes or '',
        ])
    autosize_columns(dang_vien)

    nhan_su = workbook.create_sheet('NhanSu')
    nhan_su.append(REFERENCE_HEADERS)
    for employee in employees:
        department = ' - '.join(part for part in [employee.dept_code, employee.dept_name] if part)
        nhan_su.append([
            employee.user_code,
            employee.full_name,
            employee.date_of_birth or '',
            department,
            employee.pos_name,
        ])
    autosize_columns(nhan_su)

    report = workbook.create_sheet('MapReport')
    report.append(REPORT_HEADERS)
    for row in report_rows:
        report.append(row)
    autosize_columns(report)

    manual_review = workbook.create_sheet('ManualReview')
    manual_review.append(MANUAL_REVIEW_HEADERS)
    for row in manual_review_rows:
        manual_review.append(row)
    autosize_columns(manual_review)

    workbook.save(output_path)


def write_summary(output_path: Path, summary: dict[str, object]) -> None:
    output_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')


def write_csv(output_path: Path, headers: list[str], rows: list[list[object]]) -> None:
    with output_path.open('w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def write_markdown(output_path: Path, summary: dict[str, object]) -> None:
    lines = [
        '# Party Member Bootstrap Summary',
        '',
        f"- Source rows: {summary['source_rows']}",
        f"- Employee references: {summary['employee_refs']}",
        f"- Exact name + DOB matches: {summary['exact_name_dob']}",
        f"- Name-only suggestions: {summary['name_only']}",
        f"- Auto-filled Mã NV rows: {summary['auto_filled']}",
        f"- Ambiguous matches: {summary['ambiguous']}",
        f"- Not found: {summary['not_found']}",
        '',
        'DangVien sheet already fills `Mã NV` for `EXACT_NAME_DOB` and `NAME_ONLY` rows.',
        'Review `ManualReview` and `MapReport` before importing the remaining rows.',
    ]
    output_path.write_text('\n'.join(lines) + '\n', encoding='utf-8')


def write_manual_review_markdown(output_path: Path, rows: list[list[object]]) -> None:
    lines = [
        '# Manual Review List',
        '',
        '| Dòng nguồn | Họ tên trong file | Ngày sinh | Loại đảng viên | Nhóm xử lý | Gợi ý đối chiếu | Ghi chú |',
        '| --- | --- | --- | --- | --- | --- | --- |',
    ]
    for row in rows:
        safe = [str(value or '').replace('|', '\\|') for value in row]
        lines.append(f'| {safe[0]} | {safe[1]} | {safe[2]} | {safe[3]} | {safe[4]} | {safe[5]} | {safe[6]} |')
    output_path.write_text('\n'.join(lines) + '\n', encoding='utf-8')


def main() -> int:
    parser = argparse.ArgumentParser(description='Prepare bootstrap workbook for employee party-member import.')
    parser.add_argument('--source-xlsx', required=True, help='Path to the original party-member Excel workbook.')
    parser.add_argument('--employees-tsv', required=True, help='Path to TSV export of internal_users reference data.')
    parser.add_argument('--output-dir', required=True, help='Directory to write generated artifacts into.')
    args = parser.parse_args()

    source_path = Path(args.source_xlsx).expanduser().resolve()
    employees_path = Path(args.employees_tsv).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    employees = load_employee_refs(employees_path)
    party_rows = load_party_rows(source_path)
    name_index = build_name_index(employees)

    mapping_results: list[MappingResult] = []
    report_rows: list[list[object]] = []
    manual_review_rows: list[list[object]] = []
    exact_name_dob = 0
    name_only = 0
    auto_filled = 0
    ambiguous = 0
    not_found = 0

    for row in party_rows:
        candidates = name_index.get(row.normalized_name, [])
        status, match, note = choose_match(row, candidates)
        if status == 'EXACT_NAME_DOB':
            exact_name_dob += 1
        elif status == 'NAME_ONLY':
            name_only += 1
        elif status.startswith('AMBIGUOUS'):
            ambiguous += 1
        else:
            not_found += 1

        suggestions = find_review_candidates(row, employees) if status == 'NOT_FOUND' else []
        result = MappingResult(
            party_row=row,
            status=status,
            match=match,
            note=note,
            suggestions=suggestions,
        )
        mapping_results.append(result)

        if result.auto_fill_code:
            auto_filled += 1

        department = build_department_label(match)
        position = match.pos_name if match is not None else ''

        report_rows.append([
            row.source_row,
            row.full_name,
            row.date_of_birth or '',
            result.auto_fill_code or (match.user_code if match is not None else ''),
            match.full_name if match is not None else '',
            match.date_of_birth if match is not None and match.date_of_birth else '',
            department,
            position,
            status,
            note,
        ])

        if status == 'NOT_FOUND':
            if is_note_row(row):
                review_group = 'Bỏ qua'
            else:
                review_group = 'Đối chiếu thủ công'
            manual_review_rows.append([
                row.source_row,
                row.full_name,
                row.date_of_birth or '',
                row.party_member_status,
                review_group,
                format_suggestions(suggestions),
                note,
            ])

    workbook_path = output_dir / 'party_member_bootstrap.xlsx'
    summary_path = output_dir / 'party_member_bootstrap_summary.json'
    markdown_path = output_dir / 'README.md'
    manual_review_csv_path = output_dir / 'party_member_manual_review.csv'
    manual_review_md_path = output_dir / 'party_member_manual_review.md'

    write_workbook(mapping_results, employees, workbook_path, report_rows, manual_review_rows)
    summary = {
        'source_rows': len(party_rows),
        'employee_refs': len(employees),
        'exact_name_dob': exact_name_dob,
        'name_only': name_only,
        'auto_filled': auto_filled,
        'ambiguous': ambiguous,
        'not_found': not_found,
        'generated_workbook': str(workbook_path),
        'manual_review_csv': str(manual_review_csv_path),
        'manual_review_markdown': str(manual_review_md_path),
    }
    write_summary(summary_path, summary)
    write_markdown(markdown_path, summary)
    write_csv(manual_review_csv_path, MANUAL_REVIEW_HEADERS, manual_review_rows)
    write_manual_review_markdown(manual_review_md_path, manual_review_rows)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
